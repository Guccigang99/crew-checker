import re
from datetime import datetime, timedelta
from io import BytesIO

import pandas as pd
import streamlit as st
from rapidfuzz import process, fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.comments import Comment

# =========================
# PAGE CONFIG + DESIGN
# =========================

st.set_page_config(
    page_title="FBO Hours Control tool",
    page_icon="✅",
    layout="wide",
)

st.markdown("""
<style>
    :root {
        --fbo-green: #006b3f;
        --fbo-green-dark: #004f2d;
        --fbo-green-soft: #e8f5ee;
        --fbo-border: #cfe7dc;
        --fbo-text: #10251a;
        --fbo-muted: #5d7067;
    }
    .stApp {
        background: linear-gradient(180deg, #f7fbf9 0%, #ffffff 45%, #f7fbf9 100%);
        color: var(--fbo-text);
    }
    .block-container { padding-top: 2rem; padding-bottom: 3rem; max-width: 1500px; }
    .hero {
        background: linear-gradient(135deg, #006b3f 0%, #008f55 100%);
        color: white;
        padding: 28px 32px;
        border-radius: 24px;
        margin-bottom: 24px;
        box-shadow: 0 16px 40px rgba(0, 107, 63, 0.22);
    }
    .hero h1 { margin: 0; font-size: 42px; line-height: 1.05; color: white; letter-spacing: -0.04em; }
    .hero p { margin-top: 10px; margin-bottom: 0; font-size: 16px; color: rgba(255,255,255,0.92); }
    .upload-card {
        background: white;
        border: 1px solid var(--fbo-border);
        border-radius: 18px;
        padding: 18px 20px;
        box-shadow: 0 8px 22px rgba(0, 64, 38, 0.07);
        margin-bottom: 14px;
    }
    .kpi-card {
        background: #ffffff;
        border: 1px solid var(--fbo-border);
        border-radius: 20px;
        padding: 20px 22px;
        box-shadow: 0 8px 24px rgba(0, 64, 38, 0.08);
        min-height: 126px;
    }
    .kpi-title {
        color: var(--fbo-muted);
        font-size: 14px;
        font-weight: 700;
        text-transform: uppercase;
        letter-spacing: 0.04em;
        margin-bottom: 12px;
    }
    .kpi-value { color: var(--fbo-green); font-size: 44px; font-weight: 900; line-height: 1; }
    .kpi-sub { color: var(--fbo-muted); font-size: 13px; margin-top: 8px; }
    .section-title { color: var(--fbo-green-dark); font-weight: 800; margin-top: 10px; margin-bottom: 6px; }
    div[data-testid="stFileUploader"] { background: #ffffff; border: 1px dashed #8fc7ad; border-radius: 16px; padding: 12px; }
    div[data-testid="stDownloadButton"] button, div[data-testid="stButton"] button {
        background: var(--fbo-green); color: white; border: 0; border-radius: 12px; font-weight: 800; padding: 0.65rem 1rem;
    }
    div[data-testid="stDownloadButton"] button:hover, div[data-testid="stButton"] button:hover {
        background: var(--fbo-green-dark); color: white; border: 0;
    }
    .small-note {
        background: var(--fbo-green-soft);
        border: 1px solid var(--fbo-border);
        color: var(--fbo-green-dark);
        border-radius: 14px;
        padding: 12px 14px;
        font-size: 14px;
        margin-bottom: 14px;
    }
    h2, h3 { color: var(--fbo-green-dark); }
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero">
    <h1>FBO Hours Control tool</h1>
    <p>Controleer Strobbo-weekroosters op contracturen, pauzes, rusttijden, split shifts en minderjarigenregels.</p>
</div>
""", unsafe_allow_html=True)

# =========================
# INSTELLINGEN
# =========================

MIN_DAGUREN = 2
MAX_DAGUREN_VOLWASSEN = 11
MAX_DAGUREN_MINDERJARIG = 8
MAX_WEEKUREN_VOLWASSEN = 50
MAX_WEEKUREN_MINDERJARIG = 40
MIN_RUSTUREN_TUSSEN_DAGEN = 10
MIN_SPLITSHIFT_RUSTUREN = 2
FUZZY_MATCH_SCORE = 65
MERGE_GAP_MINUTEN = 5

FILL_FOUT = PatternFill("solid", fgColor="FFFF00")
FONT_FOUT = Font(color="FF0000", bold=True)

# =========================
# HELPERS
# =========================

def normaliseer_tekst(x):
    if pd.isna(x):
        return ""
    return str(x).strip()


def normaliseer_naam(naam):
    naam = str(naam).lower().strip()
    naam = naam.replace("#", "")
    naam = re.sub(r"<\s*\d+", "", naam)
    naam = re.sub(r"\b(mgr|manager|flx|flexi|student)\b", "", naam)
    naam = naam.replace(".", "")
    naam = naam.replace("-", " ")
    naam = re.sub(r"\s+", " ", naam)
    return naam.strip()


def maak_volledige_naam(row):
    achternaam = normaliseer_tekst(row.get("NAAM", ""))
    voornaam = normaliseer_tekst(row.get("VOORNAAM", ""))
    if achternaam and voornaam:
        return f"{voornaam} {achternaam}"
    if voornaam:
        return voornaam
    if achternaam:
        return achternaam
    return ""


def bepaal_type(row):
    volledige_rij = " ".join([str(v).upper() for v in row.values])
    if "STUDENT" in volledige_rij:
        return "student"
    if "FLEXI" in volledige_rij or "FLX" in volledige_rij:
        return "flexi"
    contract = row.get("CONTRACT. UREN", row.get("CONTRACT UREN", ""))
    if pd.isna(contract) or str(contract).strip() == "":
        return "student"
    return "vast"


def veilige_float(x):
    try:
        if pd.isna(x):
            return 0
        return float(str(x).replace(",", "."))
    except Exception:
        return 0


def veilige_int(x):
    try:
        if pd.isna(x):
            return None
        leeftijd = int(float(str(x).replace(",", ".")))
        if leeftijd <= 0 or leeftijd > 100:
            return None
        return leeftijd
    except Exception:
        return None


def zoek_beste_match(strobbo_naam, crew_df):
    naam = normaliseer_naam(strobbo_naam)
    exacte_matches = []
    for _, row in crew_df.iterrows():
        voornaam = normaliseer_naam(row["VOORNAAM"])
        if naam == voornaam:
            exacte_matches.append(row["VOLLEDIGE_NAAM"])
    if len(exacte_matches) == 1:
        return exacte_matches[0], 100

    match_initiaal = re.match(r"^([a-zA-ZÀ-ÿ]+)\s+([a-zA-Z])$", naam)
    if match_initiaal:
        voornaam_gezocht = match_initiaal.group(1)
        initiaal_gezocht = match_initiaal.group(2)
        for _, row in crew_df.iterrows():
            voornaam = normaliseer_naam(row["VOORNAAM"])
            achternaam = normaliseer_naam(row["NAAM"])
            if voornaam == voornaam_gezocht and achternaam.startswith(initiaal_gezocht):
                return row["VOLLEDIGE_NAAM"], 100

    voornamen = crew_df["VOORNAAM"].astype(str).tolist()
    voornamen_norm = [normaliseer_naam(v) for v in voornamen]
    match = process.extractOne(naam, voornamen_norm, scorer=fuzz.ratio)
    if match:
        _, score, index = match
        if score >= 90:
            return crew_df.iloc[index]["VOLLEDIGE_NAAM"], round(score, 2)

    crew_namen = crew_df["VOLLEDIGE_NAAM"].tolist()
    crew_norms = [normaliseer_naam(n) for n in crew_namen]
    match = process.extractOne(naam, crew_norms, scorer=fuzz.token_sort_ratio)
    if not match:
        return None, 0
    _, score, index = match
    if score >= FUZZY_MATCH_SCORE:
        return crew_namen[index], round(score, 2)
    return None, round(score, 2)


def parse_datum(waarde):
    if isinstance(waarde, datetime):
        return waarde.date()
    tekst = str(waarde).strip().lower()
    maand_map = {"jan": 1, "feb": 2, "mrt": 3, "mar": 3, "apr": 4, "mei": 5, "jun": 6, "jul": 7, "aug": 8, "sep": 9, "okt": 10, "oct": 10, "nov": 11, "dec": 12}
    match = re.search(r"(\d{1,2})[-/\s]([a-zA-ZÀ-ÿ]+)", tekst)
    if match:
        dag = int(match.group(1))
        maand_txt = match.group(2)[:3]
        maand = maand_map.get(maand_txt)
        if maand:
            return datetime(2026, maand, dag).date()
    return None


def vind_dag_kolommen(raw):
    dag_kolommen = {}
    for rij in range(min(10, len(raw))):
        for col in range(raw.shape[1]):
            datum = parse_datum(raw.iloc[rij, col])
            if datum:
                dag_kolommen[col] = datum
        if len(dag_kolommen) >= 5:
            return dag_kolommen
    return dag_kolommen


def vind_totaal_kolom(raw):
    for rij in range(min(10, len(raw))):
        for col in range(raw.shape[1]):
            if str(raw.iloc[rij, col]).strip().lower() == "totaal":
                return col
    dag_kolommen = vind_dag_kolommen(raw)
    if dag_kolommen:
        return max(dag_kolommen.keys()) + 1
    return raw.shape[1] - 1


def parse_pauze_minuten(pauze_txt):
    if not pauze_txt:
        return 0
    match = re.search(r"(\d{1,2}):(\d{2})", str(pauze_txt))
    if not match:
        return 0
    uren = int(match.group(1))
    minuten = int(match.group(2))
    return uren * 60 + minuten


def parse_totaal_uren(totaal_txt):
    match = re.search(r"(\d{1,3}):(\d{2})", str(totaal_txt))
    if not match:
        return None
    uren = int(match.group(1))
    minuten = int(match.group(2))
    return uren + minuten / 60


def parse_shiftblokken(cell_text, datum):
    if pd.isna(cell_text) or not datum:
        return []
    tekst = str(cell_text)
    patroon = re.compile(r"(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})\s*\*?\s*(?:\n|\r|\s)*\(?\s*(\d{1,2}:\d{2})?\s*\)?", re.MULTILINE)
    blokken = []
    for match in patroon.finditer(tekst):
        start_txt = match.group(1)
        einde_txt = match.group(2)
        pauze_txt = match.group(3)
        start_dt = datetime.combine(datum, datetime.strptime(start_txt, "%H:%M").time())
        einde_dt = datetime.combine(datum, datetime.strptime(einde_txt, "%H:%M").time())
        if einde_dt <= start_dt:
            einde_dt += timedelta(days=1)
        pauze_minuten = parse_pauze_minuten(pauze_txt)
        blokken.append({"datum": datum, "start": start_dt, "einde": einde_dt, "pauze_minuten": pauze_minuten, "origineel": tekst})
    return blokken


def voeg_fout(fouten, naam, datum, fouttype, detail, ernst="Fout", cellen=None):
    fouten.append({"Medewerker": naam, "Datum": datum, "Ernst": ernst, "Fout": fouttype, "Detail": detail, "Cellen": cellen or []})


def markeer_cellen(ws, cellen, detail):
    for rij, kolom in cellen:
        try:
            cel = ws.cell(row=int(rij), column=int(kolom))
            cel.fill = FILL_FOUT
            cel.font = FONT_FOUT
            bestaande = cel.comment.text + "\n\n" if cel.comment else ""
            cel.comment = Comment(bestaande + detail, "Strobbo Checker")
        except Exception:
            pass


def show_kpi(title, value, sub=""):
    st.markdown(f"""
        <div class="kpi-card">
            <div class="kpi-title">{title}</div>
            <div class="kpi-value">{value}</div>
            <div class="kpi-sub">{sub}</div>
        </div>
        """, unsafe_allow_html=True)

# =========================
# UPLOADS
# =========================

st.markdown('<div class="section-title">Bestanden uploaden</div>', unsafe_allow_html=True)
st.markdown('<div class="small-note">Upload eerst je crew-database en daarna je Strobbo-weekrooster. De tool maakt een aparte gemarkeerde kopie, je originele bestanden worden niet aangepast.</div>', unsafe_allow_html=True)

upload_col1, upload_col2 = st.columns(2)
with upload_col1:
    st.markdown('<div class="upload-card">', unsafe_allow_html=True)
    crew_file = st.file_uploader("👥 Crew-database Excel", type=["xlsx"])
    st.markdown('</div>', unsafe_allow_html=True)
with upload_col2:
    st.markdown('<div class="upload-card">', unsafe_allow_html=True)
    rooster_file = st.file_uploader("📅 Strobbo weekrooster Excel", type=["xlsx"])
    st.markdown('</div>', unsafe_allow_html=True)

if not crew_file or not rooster_file:
    st.stop()

crew_bytes = crew_file.getvalue()
rooster_bytes = rooster_file.getvalue()

# =========================
# CREW DATABASE INLEZEN
# =========================

try:
    crew = pd.read_excel(BytesIO(crew_bytes))
except Exception as e:
    st.error(f"Kon crew-database niet lezen: {e}")
    st.stop()

crew.columns = [str(c).strip().upper() for c in crew.columns]
if "CONTRACT. UREN" not in crew.columns and "CONTRACT UREN" in crew.columns:
    crew["CONTRACT. UREN"] = crew["CONTRACT UREN"]
if "CONTRACT. UREN" not in crew.columns:
    crew["CONTRACT. UREN"] = ""
vereiste_kolommen = ["NAAM", "VOORNAAM", "LFTD"]
ontbrekend = [c for c in vereiste_kolommen if c not in crew.columns]
if ontbrekend:
    st.error(f"Deze kolommen ontbreken in je crew-database: {ontbrekend}")
    st.stop()

crew["VOLLEDIGE_NAAM"] = crew.apply(maak_volledige_naam, axis=1)
crew["TYPE"] = crew.apply(bepaal_type, axis=1)
crew["LEEFTIJD"] = crew["LFTD"].apply(veilige_int)
crew["CONTRACTUREN"] = crew["CONTRACT. UREN"].apply(veilige_float)
crew = crew[crew["VOLLEDIGE_NAAM"].str.strip() != ""].copy()

# =========================
# STROBBO ROOSTER INLEZEN
# =========================

try:
    raw = pd.read_excel(BytesIO(rooster_bytes), sheet_name=0, header=None)
    wb = load_workbook(BytesIO(rooster_bytes))
    ws = wb.active
except Exception as e:
    st.error(f"Kon Strobbo-rooster niet lezen: {e}")
    st.stop()

dag_kolommen = vind_dag_kolommen(raw)
totaal_kolom = vind_totaal_kolom(raw)
if not dag_kolommen:
    st.error("Ik kon geen datumkolommen vinden in het Strobbo-bestand.")
    st.stop()

# =========================
# SHIFTS EXTRACTEN + TOTAALCELLEN BIJHOUDEN
# =========================

shifts = []
totaal_info = {}
huidige_medewerker = None
for rij_index, row in raw.iterrows():
    eerste_cel = normaliseer_tekst(row.iloc[0])
    if eerste_cel.startswith("#"):
        huidige_medewerker = eerste_cel.replace("#", "").strip()
    elif eerste_cel and not parse_datum(eerste_cel):
        totaal_txt = row.iloc[totaal_kolom] if totaal_kolom < len(row) else ""
        if parse_totaal_uren(totaal_txt) is not None:
            huidige_medewerker = eerste_cel.strip()
    if not huidige_medewerker:
        continue

    if totaal_kolom < len(row):
        totaal_txt = row.iloc[totaal_kolom]
        totaal_uren = parse_totaal_uren(totaal_txt)
        if totaal_uren is not None:
            totaal_info[normaliseer_naam(huidige_medewerker)] = {"uren": totaal_uren, "cel": (rij_index + 1, totaal_kolom + 1), "tekst": str(totaal_txt)}

    for col_index, datum in dag_kolommen.items():
        if col_index >= len(row):
            continue
        cel = row.iloc[col_index]
        gevonden_shifts = parse_shiftblokken(cel, datum)
        for shift in gevonden_shifts:
            shift["strobbo_naam"] = huidige_medewerker
            shift["bron_cellen"] = [(rij_index + 1, col_index + 1)]
            shifts.append(shift)

shifts_df = pd.DataFrame(shifts)
if shifts_df.empty:
    st.warning("Geen shifts gevonden in het Strobbo-bestand.")
    st.stop()

# =========================
# SHIFTS SAMENVOEGEN OVER RIJEN
# =========================

shifts_df = shifts_df.sort_values(["strobbo_naam", "datum", "start", "einde"]).reset_index(drop=True)
samengevoegde_shifts = []
for (naam, datum), groep in shifts_df.groupby(["strobbo_naam", "datum"]):
    groep = groep.sort_values("start").reset_index(drop=True)
    huidige = groep.iloc[0].to_dict()
    for i in range(1, len(groep)):
        blok = groep.iloc[i].to_dict()
        gap_minuten = (blok["start"] - huidige["einde"]).total_seconds() / 60
        if gap_minuten <= MERGE_GAP_MINUTEN:
            huidige["einde"] = max(huidige["einde"], blok["einde"])
            huidige["pauze_minuten"] += blok["pauze_minuten"]
            huidige["bron_cellen"] = list(huidige.get("bron_cellen", [])) + list(blok.get("bron_cellen", []))
            huidige["origineel"] = str(huidige.get("origineel", "")) + "\\n---\\n" + str(blok.get("origineel", ""))
        else:
            bruto_uren = (huidige["einde"] - huidige["start"]).total_seconds() / 3600
            huidige["bruto_uren"] = bruto_uren
            huidige["netto_uren"] = bruto_uren - (huidige["pauze_minuten"] / 60)
            samengevoegde_shifts.append(huidige)
            huidige = blok
    bruto_uren = (huidige["einde"] - huidige["start"]).total_seconds() / 3600
    huidige["bruto_uren"] = bruto_uren
    huidige["netto_uren"] = bruto_uren - (huidige["pauze_minuten"] / 60)
    samengevoegde_shifts.append(huidige)
shifts_df = pd.DataFrame(samengevoegde_shifts)

# =========================
# MATCHEN MET DATABASE
# =========================

match_resultaten = []
for naam in shifts_df["strobbo_naam"].unique():
    beste_naam, score = zoek_beste_match(naam, crew)
    match_resultaten.append({"Strobbo naam": naam, "Database naam": beste_naam if beste_naam else "NIET GEVONDEN", "Match score": round(score, 2)})
match_df = pd.DataFrame(match_resultaten)
match_map = {row["Strobbo naam"]: row["Database naam"] for _, row in match_df.iterrows() if row["Database naam"] != "NIET GEVONDEN"}
shifts_df["database_naam"] = shifts_df["strobbo_naam"].map(match_map)

# =========================
# FOUTEN CONTROLEREN
# =========================

fouten = []
for _, row in match_df.iterrows():
    if row["Database naam"] == "NIET GEVONDEN":
        voeg_fout(fouten, row["Strobbo naam"], "", "Naam niet gevonden", f"Geen goede match in database. Match score: {row['Match score']}", "Waarschuwing", [])

for _, shift in shifts_df.dropna(subset=["database_naam"]).iterrows():
    naam = shift["database_naam"]
    datum = shift["datum"]
    cellen = shift.get("bron_cellen", [])
    persoon = crew[crew["VOLLEDIGE_NAAM"] == naam].iloc[0]
    leeftijd = persoon["LEEFTIJD"]
    netto_uren = shift["netto_uren"]
    bruto_uren = shift["bruto_uren"]
    pauze = shift["pauze_minuten"]

    if netto_uren < MIN_DAGUREN:
        voeg_fout(fouten, naam, datum, "Shift te kort", f"{netto_uren:.2f}u gewerkt. Minimum is {MIN_DAGUREN}u.", "Fout", cellen)
    if leeftijd is not None and leeftijd < 18:
        if netto_uren > MAX_DAGUREN_MINDERJARIG:
            voeg_fout(fouten, naam, datum, "Minderjarige werkt te lang", f"{netto_uren:.2f}u gewerkt. Maximum voor <18 is {MAX_DAGUREN_MINDERJARIG}u.", "Fout", cellen)
    else:
        if netto_uren > MAX_DAGUREN_VOLWASSEN:
            voeg_fout(fouten, naam, datum, "Shift te lang", f"{netto_uren:.2f}u gewerkt. Maximum is {MAX_DAGUREN_VOLWASSEN}u.", "Fout", cellen)

    if leeftijd is None or leeftijd >= 18:
        if bruto_uren <= 5 and pauze > 0:
            voeg_fout(fouten, naam, datum, "Onnodige pauze", f"{bruto_uren:.2f}u shift heeft {pauze} min pauze, maar tot 5u is geen pauze nodig.", "Waarschuwing", cellen)
        elif 5 < bruto_uren <= 8 and pauze < 20:
            voeg_fout(fouten, naam, datum, "Pauze ontbreekt", f"{bruto_uren:.2f}u shift. Minstens 20 min pauze nodig. Geplande pauze: {pauze} min.", "Fout", cellen)
        elif bruto_uren > 8 and pauze < 30:
            voeg_fout(fouten, naam, datum, "Pauze ontbreekt", f"{bruto_uren:.2f}u shift. Minstens 30 min pauze nodig. Geplande pauze: {pauze} min.", "Fout", cellen)

    if leeftijd is not None and leeftijd < 18:
        if bruto_uren > 4.5 and bruto_uren <= 6 and pauze < 30:
            voeg_fout(fouten, naam, datum, "Pauze minderjarige ontbreekt", f"{bruto_uren:.2f}u shift. Minderjarige heeft minstens 30 min pauze nodig. Geplande pauze: {pauze} min.", "Fout", cellen)
        elif bruto_uren > 6 and pauze < 60:
            voeg_fout(fouten, naam, datum, "Pauze minderjarige ontbreekt", f"{bruto_uren:.2f}u shift. Minderjarige heeft minstens 60 min pauze nodig. Geplande pauze: {pauze} min.", "Fout", cellen)

    if leeftijd is not None:
        einduur = shift["einde"].hour + shift["einde"].minute / 60
        if leeftijd <= 15 and einduur > 20:
            voeg_fout(fouten, naam, datum, "Nachtwerk minderjarige", f"{leeftijd} jaar en werkt tot {shift['einde'].strftime('%H:%M')}. Max tot 20:00.", "Fout", cellen)
        elif 16 <= leeftijd < 18 and einduur > 23:
            voeg_fout(fouten, naam, datum, "Nachtwerk minderjarige", f"{leeftijd} jaar en werkt tot {shift['einde'].strftime('%H:%M')}. Max tot 23:00.", "Fout", cellen)

# Weekuren + contracturen op basis van Strobbo totaalcel
for naam, groep in shifts_df.dropna(subset=["database_naam"]).groupby("database_naam"):
    persoon = crew[crew["VOLLEDIGE_NAAM"] == naam].iloc[0]
    leeftijd = persoon["LEEFTIJD"]
    medewerker_type = persoon["TYPE"]
    contracturen = persoon["CONTRACTUREN"]
    strobbo_naam = groep.iloc[0]["strobbo_naam"]
    info = totaal_info.get(normaliseer_naam(strobbo_naam), {})
    totaal_weekuren = info.get("uren", groep["netto_uren"].sum())
    totaal_cel = [info["cel"]] if "cel" in info else []
    if leeftijd is not None and leeftijd < 18:
        if totaal_weekuren > MAX_WEEKUREN_MINDERJARIG:
            voeg_fout(fouten, naam, "", "Weekuren overschreden", f"{totaal_weekuren:.2f}u gepland volgens Strobbo. Maximum voor <18 is {MAX_WEEKUREN_MINDERJARIG}u.", "Fout", totaal_cel)
    else:
        if totaal_weekuren > MAX_WEEKUREN_VOLWASSEN:
            voeg_fout(fouten, naam, "", "Weekuren overschreden", f"{totaal_weekuren:.2f}u gepland volgens Strobbo. Maximum is {MAX_WEEKUREN_VOLWASSEN}u.", "Fout", totaal_cel)
    if medewerker_type == "vast" and contracturen > 0 and totaal_weekuren < contracturen:
        tekort = contracturen - totaal_weekuren
        voeg_fout(fouten, naam, "", "Contracturen niet gehaald", f"{totaal_weekuren:.2f}u gepland volgens Strobbo, contract is {contracturen:.2f}u. Tekort: {tekort:.2f}u.", "Fout", totaal_cel)

# Rust tussen shifts
# - zelfde Strobbo-dag = split shift, minimum 2u
# - andere dag = minimum 10u rust tussen dagen
for naam, groep in shifts_df.dropna(subset=["database_naam"]).groupby("database_naam"):
    groep = groep.sort_values("start").reset_index(drop=True)
    for i in range(1, len(groep)):
        vorige = groep.iloc[i - 1]
        huidige = groep.iloc[i]
        rusturen = (huidige["start"] - vorige["einde"]).total_seconds() / 3600
        if vorige["datum"] == huidige["datum"]:
            if rusturen < MIN_SPLITSHIFT_RUSTUREN:
                voeg_fout(fouten, naam, huidige["datum"], "Split shift te kort", f"Slechts {rusturen:.2f}u tussen 2 shifts op dezelfde dag. Minimum is {MIN_SPLITSHIFT_RUSTUREN}u.", "Fout", huidige.get("bron_cellen", []))
        else:
            if rusturen < MIN_RUSTUREN_TUSSEN_DAGEN:
                voeg_fout(fouten, naam, huidige["datum"], "Te weinig rust tussen dagen", f"Slechts {rusturen:.2f}u rust tussen {vorige['einde'].strftime('%d/%m %H:%M')} en {huidige['start'].strftime('%d/%m %H:%M')}. Minimum is {MIN_RUSTUREN_TUSSEN_DAGEN}u.", "Fout", huidige.get("bron_cellen", []))

fouten_df = pd.DataFrame(fouten)

# =========================
# EXCEL MARKEREN
# =========================

if not fouten_df.empty:
    for _, fout in fouten_df.iterrows():
        markeer_cellen(ws, fout.get("Cellen", []), f"{fout['Fout']}: {fout['Detail']}")

excel_buffer = BytesIO()
wb.save(excel_buffer)

# =========================
# OUTPUT
# =========================

aantal_fouten = 0 if fouten_df.empty else len(fouten_df[fouten_df["Ernst"] == "Fout"])
aantal_waarschuwingen = 0 if fouten_df.empty else len(fouten_df[fouten_df["Ernst"] == "Waarschuwing"])
aantal_shifts = len(shifts_df)
aantal_medewerkers = shifts_df["database_naam"].dropna().nunique()

kpi1, kpi2, kpi3, kpi4 = st.columns(4)
with kpi1:
    show_kpi("Fouten", aantal_fouten, "moeten gecorrigeerd worden")
with kpi2:
    show_kpi("Waarschuwingen", aantal_waarschuwingen, "nakijken indien nodig")
with kpi3:
    show_kpi("Shifts", aantal_shifts, "gevonden na samenvoegen")
with kpi4:
    show_kpi("Medewerkers", aantal_medewerkers, "gematcht met database")

tab_overzicht, tab_fouten, tab_shifts, tab_match, tab_export = st.tabs(["📊 Overzicht", "🚨 Fouten", "🕒 Shifts", "🔗 Naamkoppeling", "📥 Export"])

with tab_overzicht:
    st.subheader("📊 Weekuren per medewerker")
    weekuren = []
    for naam, groep in shifts_df.dropna(subset=["database_naam"]).groupby("database_naam"):
        strobbo_naam = groep.iloc[0]["strobbo_naam"]
        info = totaal_info.get(normaliseer_naam(strobbo_naam), {})
        totaal = info.get("uren", groep["netto_uren"].sum())
        persoon = crew[crew["VOLLEDIGE_NAAM"] == naam].iloc[0]
        weekuren.append({"Medewerker": naam, "Weekuren": round(totaal, 2), "Type": persoon["TYPE"], "Leeftijd": persoon["LEEFTIJD"], "Contracturen": persoon["CONTRACTUREN"]})
    st.dataframe(pd.DataFrame(weekuren), use_container_width=True)

with tab_fouten:
    st.subheader("🚨 Foutenrapport")
    if fouten_df.empty:
        st.success("Geen fouten gevonden volgens de ingestelde regels.")
    else:
        st.error(f"{len(fouten_df)} fout(en) of waarschuwing(en) gevonden.")
        st.dataframe(fouten_df.drop(columns=["Cellen"], errors="ignore"), use_container_width=True)

with tab_shifts:
    st.subheader("📋 Gevonden shifts")
    toon_shifts = shifts_df.copy()
    toon_shifts["Start"] = toon_shifts["start"].dt.strftime("%d/%m/%Y %H:%M")
    toon_shifts["Einde"] = toon_shifts["einde"].dt.strftime("%d/%m/%Y %H:%M")
    toon_shifts["Netto uren"] = toon_shifts["netto_uren"].round(2)
    toon_shifts["Bruto uren"] = toon_shifts["bruto_uren"].round(2)
    st.dataframe(toon_shifts[["strobbo_naam", "database_naam", "datum", "Start", "Einde", "pauze_minuten", "Bruto uren", "Netto uren"]], use_container_width=True)

with tab_match:
    st.subheader("✅ Naamkoppeling Strobbo ↔ Database")
    st.dataframe(match_df, use_container_width=True)

with tab_export:
    st.subheader("📥 Downloads")
    st.write("Download hier de gemarkeerde Strobbo Excel of het losse foutenrapport.")
    st.download_button(label="📥 Download gemarkeerde Strobbo Excel", data=excel_buffer.getvalue(), file_name="strobbo_rooster_gecontroleerd.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    rapport_buffer = BytesIO()
    with pd.ExcelWriter(rapport_buffer, engine="openpyxl") as writer:
        fouten_df.drop(columns=["Cellen"], errors="ignore").to_excel(writer, index=False, sheet_name="Foutenrapport")
        match_df.to_excel(writer, index=False, sheet_name="Naamkoppeling")
        toon_export = shifts_df.copy()
        toon_export["Start"] = toon_export["start"].dt.strftime("%d/%m/%Y %H:%M")
        toon_export["Einde"] = toon_export["einde"].dt.strftime("%d/%m/%Y %H:%M")
        toon_export.to_excel(writer, index=False, sheet_name="Gevonden shifts")
    st.download_button(label="📥 Download foutenrapport als Excel", data=rapport_buffer.getvalue(), file_name="foutenrapport_strobbo.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
